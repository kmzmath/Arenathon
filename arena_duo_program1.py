#!/usr/bin/env python3
"""
Programa 1 — coletor/validador de partidas de duplas para campeonato de Arena.

O script:
1) Lê as duplas de um arquivo XLSX com colunas:
   - duo_id
   - player1_riot_id
   - player2_riot_id
   - duo_name
2) Resolve todos os Riot IDs para PUUID/plataforma/cluster.
3) Busca partidas Arena (queues 1700 e 1710) apenas na janela do campeonato.
4) Baixa o JSON bruto de cada partida e, opcionalmente, a timeline.
5) Valida se a dupla realmente jogou junta na mesma subteam da partida.
6) Calcula a pontuação por colocação da dupla.
7) Gera arquivos para persistência, auditoria e uso futuro no front-end.

Uso básico:
    python arena_duo_program1.py \
        --duos-xlsx ./duplas.xlsx \
        --api-key SUA_CHAVE \
        --output ./arena_program1_dump

Observações:
- A janela default é a do campeonato:
  2026-04-29 13:00:00 até 2026-05-13 23:59:59 em America/Sao_Paulo.
- O script revalida a janela no payload bruto da partida, mesmo usando startTime/endTime na API.
- Uma partida válida é contabilizada por "dupla + match_id", nunca por jogador.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import random
import re
import sys
import threading
import time
from collections import deque
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Deque, Dict, Iterable, List, Optional, Set, Tuple
from urllib.parse import quote
from zoneinfo import ZoneInfo

import openpyxl
import requests


ACCOUNT_CLUSTER = "americas"
ARENA_QUEUES = (1700, 1710)
TOURNAMENT_TZ = ZoneInfo("America/Sao_Paulo")
POINTS_BY_PLACEMENT = {
    1: 10,
    2: 6,
    3: 3,
    4: 1,
    5: -1,
    6: -3,
    7: -5,
    8: -8,
}

DEFAULT_START_LOCAL = datetime(2026, 1, 29, 13, 0, 0, tzinfo=TOURNAMENT_TZ)
DEFAULT_END_LOCAL = datetime(2026, 5, 13, 23, 59, 59, tzinfo=TOURNAMENT_TZ)

# Janela oficial
#DEFAULT_START_LOCAL = datetime(2026, 4, 29, 13, 0, 0, tzinfo=TOURNAMENT_TZ)
#DEFAULT_END_LOCAL = datetime(2026, 5, 13, 23, 59, 59, tzinfo=TOURNAMENT_TZ)

PLATFORM_TO_REGION = {
    "BR1": "americas",
    "LA1": "americas",
    "LA2": "americas",
    "NA1": "americas",
    "EUN1": "europe",
    "EUW1": "europe",
    "RU": "europe",
    "TR1": "europe",
    "JP1": "asia",
    "KR": "asia",
    "OC1": "sea",
    "SG2": "sea",
    "TW2": "sea",
    "VN2": "sea",
    "PH2": "sea",
    "TH2": "sea",
}

DEFAULT_TIMEOUT = 30
MAX_RETRIES = 8


class RiotAPIError(RuntimeError):
    """Erro de API Riot com contexto adicional."""


@dataclass
class TournamentWindow:
    start_local_iso: str
    end_local_iso: str
    start_epoch_seconds: int
    end_epoch_seconds: int


@dataclass
class PlayerRef:
    riot_id: str
    game_name: str
    tag_line: str
    puuid: str
    platform: str
    regional_cluster: str
    slug: str


@dataclass
class DuoRef:
    duo_id: str
    duo_name: str
    player1_riot_id: str
    player2_riot_id: str
    player1_puuid: str
    player2_puuid: str
    platform: str
    regional_cluster: str
    slug: str


def default_window() -> TournamentWindow:
    return TournamentWindow(
        start_local_iso=DEFAULT_START_LOCAL.isoformat(),
        end_local_iso=DEFAULT_END_LOCAL.isoformat(),
        start_epoch_seconds=int(DEFAULT_START_LOCAL.timestamp()),
        end_epoch_seconds=int(DEFAULT_END_LOCAL.timestamp()),
    )


def slugify(value: str) -> str:
    value = str(value).strip().replace("#", "_")
    value = re.sub(r"[^\w\-.]+", "_", value, flags=re.UNICODE)
    value = re.sub(r"_+", "_", value)
    return value.strip("._") or "item"


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def save_json(path: Path, data: Any) -> None:
    """Escreve JSON via arquivo temporário e troca atômica.

    Isso impede que a API leia um JSON parcialmente escrito enquanto o
    coletor está gerando uma nova rodada de dados.
    """
    ensure_dir(path.parent)
    content = json.dumps(data, ensure_ascii=False, indent=2)
    tmp_path = path.with_name(f".{path.name}.{os.getpid()}.{threading.get_ident()}.tmp")

    try:
        with tmp_path.open("w", encoding="utf-8") as fp:
            fp.write(content)
            fp.flush()
            os.fsync(fp.fileno())
        os.replace(tmp_path, path)
    finally:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception:
            logging.exception("Falha removendo arquivo temporário %s", tmp_path)


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def split_riot_id(riot_id: str) -> Tuple[str, str]:
    riot_id = str(riot_id).strip()
    if "#" not in riot_id:
        raise ValueError(f'Riot ID inválido: "{riot_id}". Use o formato gameName#tagLine.')
    game_name, tag_line = riot_id.rsplit("#", 1)
    game_name = game_name.strip()
    tag_line = tag_line.strip()
    if not game_name or not tag_line:
        raise ValueError(f'Riot ID inválido: "{riot_id}".')
    return game_name, tag_line


def dedupe_keep_order(values: Iterable[str]) -> List[str]:
    seen: Set[str] = set()
    result: List[str] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def platform_to_regional_cluster(platform: str) -> str:
    platform = str(platform).upper()
    try:
        return PLATFORM_TO_REGION[platform]
    except KeyError as exc:
        known = ", ".join(sorted(PLATFORM_TO_REGION))
        raise ValueError(
            f"Plataforma não mapeada: {platform}. Ajuste PLATFORM_TO_REGION. Conhecidas: {known}"
        ) from exc


def normalize_game_duration_seconds(info: Dict[str, Any]) -> Optional[float]:
    game_duration = info.get("gameDuration")
    if game_duration is None:
        return None
    if info.get("gameEndTimestamp") is not None:
        return float(game_duration)
    return float(game_duration) / 1000.0


def epoch_ms_to_local_iso(epoch_ms: Optional[int]) -> Optional[str]:
    if epoch_ms is None:
        return None
    return datetime.fromtimestamp(epoch_ms / 1000, tz=TOURNAMENT_TZ).isoformat()


def points_from_placement(placement: Optional[int]) -> Optional[int]:
    if placement is None:
        return None
    return POINTS_BY_PLACEMENT.get(int(placement))


def flatten_csv_rows(path: Path, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        return

    fieldnames: List[str] = []
    seen: Set[str] = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                fieldnames.append(key)

    ensure_dir(path.parent)
    with path.open("w", newline="", encoding="utf-8") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def load_duos_xlsx(path: Path) -> List[Dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("A planilha de duplas está vazia.")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    required = ["duo_id", "player1_riot_id", "player2_riot_id", "duo_name"]
    missing = [col for col in required if col not in header]
    if missing:
        raise ValueError(f"Colunas obrigatórias ausentes no XLSX: {missing}. Cabeçalho encontrado: {header}")

    index = {name: header.index(name) for name in required}
    result: List[Dict[str, str]] = []

    for line_no, row in enumerate(rows[1:], start=2):
        if row is None:
            continue

        duo_id = row[index["duo_id"]]
        player1 = row[index["player1_riot_id"]]
        player2 = row[index["player2_riot_id"]]
        duo_name = row[index["duo_name"]]

        if duo_id is None and player1 is None and player2 is None and duo_name is None:
            continue

        duo_id_str = str(duo_id).strip() if duo_id is not None else ""
        player1_str = str(player1).strip() if player1 is not None else ""
        player2_str = str(player2).strip() if player2 is not None else ""
        duo_name_str = str(duo_name).strip() if duo_name is not None else ""

        if not duo_id_str:
            raise ValueError(f"Linha {line_no}: duo_id vazio.")
        if not player1_str:
            raise ValueError(f"Linha {line_no}: player1_riot_id vazio.")
        if not player2_str:
            raise ValueError(f"Linha {line_no}: player2_riot_id vazio.")
        if not duo_name_str:
            raise ValueError(f"Linha {line_no}: duo_name vazio.")

        result.append({
            "duo_id": duo_id_str,
            "player1_riot_id": player1_str,
            "player2_riot_id": player2_str,
            "duo_name": duo_name_str,
        })

    if not result:
        raise ValueError("Nenhuma dupla válida encontrada na planilha.")
    return result


def load_players_cache(path: Path) -> Dict[str, PlayerRef]:
    """
    Carrega `players_resolved.json` gerado por uma execução anterior e devolve
    um dict {riot_id: PlayerRef} só com entradas válidas (sem erros).

    Riot IDs não mudam de PUUID durante um campeonato, então re-resolver todos
    a cada ciclo é desperdício de chamadas de API. Quando um riot_id novo
    aparecer na planilha, ele simplesmente não estará no cache e será
    resolvido pelo fluxo normal.
    """
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        logging.exception("Falha lendo cache de players em %s. Ignorando.", path)
        return {}

    cached: Dict[str, PlayerRef] = {}
    required_fields = ("riot_id", "game_name", "tag_line", "puuid", "platform", "regional_cluster", "slug")
    for entry in data.get("players", []):
        if not isinstance(entry, dict):
            continue
        if any(not entry.get(f) for f in required_fields):
            continue
        try:
            cached[entry["riot_id"]] = PlayerRef(
                riot_id=entry["riot_id"],
                game_name=entry["game_name"],
                tag_line=entry["tag_line"],
                puuid=entry["puuid"],
                platform=entry["platform"],
                regional_cluster=entry["regional_cluster"],
                slug=entry["slug"],
            )
        except Exception:
            logging.exception("Entrada inválida no cache de players: %r", entry)
    return cached


class RateLimiter:
    """
    Sliding-window rate limiter para respeitar os limites da Riot Dev API:
      - 20 requests por 1 segundo
      - 100 requests por 2 minutos

    Usa margens de segurança (18/s e 95/2min por default) para evitar
    flutuações de clock e contagem do lado do servidor.

    Cada chamada a acquire() bloqueia o tempo necessário e só retorna
    quando for seguro disparar uma nova requisição. Thread-safe.
    """

    def __init__(
        self,
        short_limit: int = 18,
        short_period: float = 1.0,
        long_limit: int = 95,
        long_period: float = 120.0,
    ) -> None:
        self.short_limit = short_limit
        self.short_period = short_period
        self.long_limit = long_limit
        self.long_period = long_period
        self._timestamps: Deque[float] = deque()
        self._lock = threading.Lock()

    def acquire(self) -> None:
        while True:
            with self._lock:
                now = time.monotonic()

                # Remove timestamps fora da janela longa
                while self._timestamps and now - self._timestamps[0] > self.long_period:
                    self._timestamps.popleft()

                count_long = len(self._timestamps)
                count_short = sum(1 for t in self._timestamps if now - t <= self.short_period)

                wait = 0.0
                if count_long >= self.long_limit:
                    wait = max(wait, self.long_period - (now - self._timestamps[0]) + 0.05)
                if count_short >= self.short_limit:
                    # Primeiro timestamp dentro da janela curta (a deque é cronológica)
                    for t in self._timestamps:
                        if now - t <= self.short_period:
                            wait = max(wait, self.short_period - (now - t) + 0.05)
                            break

                if wait <= 0:
                    self._timestamps.append(now)
                    return

            # Libera o lock durante o sleep para não segurar outras threads à toa.
            # Ao acordar, relê o estado dentro do lock antes de decidir.
            logging.info(
                "Rate limiter: aguardando %.2fs (janela longa=%d/%d, curta=%d/%d)",
                wait, count_long, self.long_limit, count_short, self.short_limit,
            )
            time.sleep(wait)


class RiotClient:
    def __init__(
        self,
        api_key: str,
        timeout: int = DEFAULT_TIMEOUT,
        rate_limiter: Optional[RateLimiter] = None,
    ) -> None:
        self.timeout = timeout
        self.rate_limiter = rate_limiter or RateLimiter()
        self.session = requests.Session()
        self.session.headers.update({
            "X-Riot-Token": api_key,
            "Accept": "application/json",
            "User-Agent": "arena-duo-program1/1.0",
        })

    def _request_json(self, url: str, params: Optional[Dict[str, Any]] = None) -> Any:
        attempt = 0
        while True:
            attempt += 1
            # Gatekeeper: garante que nunca passamos dos limites da Riot,
            # mesmo contando retries de 429 / 5xx.
            self.rate_limiter.acquire()
            try:
                response = self.session.get(url, params=params, timeout=self.timeout)
            except requests.RequestException as exc:
                if attempt >= MAX_RETRIES:
                    raise RiotAPIError(f"Falha de rede após {attempt} tentativas: {url}") from exc
                sleep_s = min(2 ** attempt, 30) + random.random()
                logging.warning("Erro de rede em %s. Tentando novamente em %.1fs", url, sleep_s)
                time.sleep(sleep_s)
                continue

            if response.status_code == 200:
                return response.json()

            if response.status_code == 404:
                raise RiotAPIError(f"Recurso não encontrado (404): {url}")

            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After")
                sleep_s = float(retry_after) if retry_after else min(2 ** attempt, 60)
                sleep_s += random.random()
                logging.warning("Rate limit em %s. Aguardando %.1fs", url, sleep_s)
                time.sleep(sleep_s)
                if attempt >= MAX_RETRIES:
                    raise RiotAPIError(f"Rate limit persistente após {attempt} tentativas: {url}")
                continue

            if response.status_code in {500, 502, 503, 504}:
                if attempt >= MAX_RETRIES:
                    raise RiotAPIError(
                        f"Erro {response.status_code} após {attempt} tentativas: {url}\n{response.text}"
                    )
                sleep_s = min(2 ** attempt, 60) + random.random()
                logging.warning("Erro %s em %s. Tentando novamente em %.1fs", response.status_code, url, sleep_s)
                time.sleep(sleep_s)
                continue

            raise RiotAPIError(
                f"Erro HTTP {response.status_code} em {url}\nResposta: {response.text}"
            )

    def resolve_account_by_riot_id(self, game_name: str, tag_line: str) -> Dict[str, Any]:
        url = (
            f"https://{ACCOUNT_CLUSTER}.api.riotgames.com"
            f"/riot/account/v1/accounts/by-riot-id/{quote(game_name)}/{quote(tag_line)}"
        )
        return self._request_json(url)

    def get_active_platform(self, puuid: str) -> str:
        url = (
            f"https://{ACCOUNT_CLUSTER}.api.riotgames.com"
            f"/riot/account/v1/region/by-game/lol/by-puuid/{quote(puuid)}"
        )
        payload = self._request_json(url)
        platform = str(payload.get("region", "")).upper()
        if not platform:
            raise RiotAPIError(f"Não foi possível descobrir a plataforma para o PUUID {puuid}")
        return platform

    def get_match_ids(
        self,
        regional_cluster: str,
        puuid: str,
        queue_id: int,
        start_time: Optional[int] = None,
        end_time: Optional[int] = None,
        max_matches: Optional[int] = None,
    ) -> List[str]:
        base_url = (
            f"https://{regional_cluster}.api.riotgames.com"
            f"/lol/match/v5/matches/by-puuid/{quote(puuid)}/ids"
        )

        start = 0
        collected: List[str] = []

        while True:
            remaining = None if max_matches is None else max_matches - len(collected)
            if remaining is not None and remaining <= 0:
                break

            count = 100 if remaining is None else max(0, min(100, remaining))
            params: Dict[str, Any] = {
                "queue": queue_id,
                "start": start,
                "count": count,
            }
            if start_time is not None:
                params["startTime"] = start_time
            if end_time is not None:
                params["endTime"] = end_time

            batch = self._request_json(base_url, params=params)
            if not isinstance(batch, list):
                raise RiotAPIError(f"Resposta inesperada em lista de partidas: {base_url} -> {batch!r}")

            if not batch:
                break

            collected.extend(batch)
            logging.info(
                "Queue %s | player=%s | start=%s | recebidos=%s | total=%s",
                queue_id, puuid, start, len(batch), len(collected)
            )

            if len(batch) < count:
                break

            start += len(batch)

        return collected

    def get_match(self, regional_cluster: str, match_id: str) -> Dict[str, Any]:
        url = f"https://{regional_cluster}.api.riotgames.com/lol/match/v5/matches/{quote(match_id)}"
        return self._request_json(url)

    def get_timeline(self, regional_cluster: str, match_id: str) -> Dict[str, Any]:
        url = f"https://{regional_cluster}.api.riotgames.com/lol/match/v5/matches/{quote(match_id)}/timeline"
        return self._request_json(url)


def find_participant(match_payload: Dict[str, Any], puuid: str) -> Optional[Dict[str, Any]]:
    for participant in match_payload.get("info", {}).get("participants", []):
        if participant.get("puuid") == puuid:
            return participant
    return None


def summarize_participant(participant: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "puuid": participant.get("puuid"),
        "riot_id_game_name": participant.get("riotIdGameName"),
        "riot_id_tagline": participant.get("riotIdTagline"),
        "summoner_name": participant.get("summonerName"),
        "participant_id": participant.get("participantId"),
        "team_id": participant.get("teamId"),
        "player_subteam_id": participant.get("playerSubteamId"),
        "subteam_placement": participant.get("subteamPlacement"),
        "champion_id": participant.get("championId"),
        "champion_name": participant.get("championName"),
        "champ_level": participant.get("champLevel"),
        "placement": participant.get("placement"),
        "kills": participant.get("kills"),
        "deaths": participant.get("deaths"),
        "assists": participant.get("assists"),
        "gold_earned": participant.get("goldEarned"),
        "gold_spent": participant.get("goldSpent"),
        "total_damage_dealt_to_champions": participant.get("totalDamageDealtToChampions"),
        "total_damage_taken": participant.get("totalDamageTaken"),
        "damage_self_mitigated": participant.get("damageSelfMitigated"),
        "vision_score": participant.get("visionScore"),
        "time_played": participant.get("timePlayed"),
        "item0": participant.get("item0"),
        "item1": participant.get("item1"),
        "item2": participant.get("item2"),
        "item3": participant.get("item3"),
        "item4": participant.get("item4"),
        "item5": participant.get("item5"),
        "item6": participant.get("item6"),
        "player_augment1": participant.get("playerAugment1"),
        "player_augment2": participant.get("playerAugment2"),
        "player_augment3": participant.get("playerAugment3"),
        "player_augment4": participant.get("playerAugment4"),
        "win": participant.get("win"),
    }


def build_valid_row(
    duo: DuoRef,
    match_payload: Dict[str, Any],
    timeline_path: Optional[Path],
    match_path: Path,
    source_label: str,
) -> Dict[str, Any]:
    info = match_payload.get("info", {})
    metadata = match_payload.get("metadata", {})

    p1 = find_participant(match_payload, duo.player1_puuid)
    p2 = find_participant(match_payload, duo.player2_puuid)
    if p1 is None or p2 is None:
        raise ValueError("build_valid_row chamado sem os dois participantes presentes.")

    placement = p1.get("subteamPlacement")
    points = points_from_placement(placement)

    return {
        "duo_id": duo.duo_id,
        "duo_name": duo.duo_name,
        "player1_riot_id": duo.player1_riot_id,
        "player2_riot_id": duo.player2_riot_id,
        "player1_puuid": duo.player1_puuid,
        "player2_puuid": duo.player2_puuid,
        "platform": duo.platform,
        "regional_cluster": duo.regional_cluster,
        "match_id": metadata.get("matchId"),
        "queue_id": info.get("queueId"),
        "map_id": info.get("mapId"),
        "game_mode": info.get("gameMode"),
        "game_type": info.get("gameType"),
        "game_version": info.get("gameVersion"),
        "game_creation": info.get("gameCreation"),
        "game_start_timestamp": info.get("gameStartTimestamp"),
        "game_end_timestamp": info.get("gameEndTimestamp"),
        "game_start_local": epoch_ms_to_local_iso(info.get("gameStartTimestamp")),
        "game_end_local": epoch_ms_to_local_iso(info.get("gameEndTimestamp")),
        "game_duration_seconds": normalize_game_duration_seconds(info),
        "end_of_game_result": info.get("endOfGameResult"),
        "player_subteam_id": p1.get("playerSubteamId"),
        "subteam_placement": placement,
        "points": points,
        "candidate_source": source_label,
        "match_json_path": str(match_path.resolve()),
        "timeline_json_path": str(timeline_path.resolve()) if timeline_path else None,
        "player1": summarize_participant(p1),
        "player2": summarize_participant(p2),
    }


def build_valid_csv_row(valid_row: Dict[str, Any]) -> Dict[str, Any]:
    p1 = valid_row["player1"]
    p2 = valid_row["player2"]

    return {
        "duo_id": valid_row["duo_id"],
        "duo_name": valid_row["duo_name"],
        "player1_riot_id": valid_row["player1_riot_id"],
        "player2_riot_id": valid_row["player2_riot_id"],
        "player1_puuid": valid_row["player1_puuid"],
        "player2_puuid": valid_row["player2_puuid"],
        "platform": valid_row["platform"],
        "regional_cluster": valid_row["regional_cluster"],
        "match_id": valid_row["match_id"],
        "queue_id": valid_row["queue_id"],
        "map_id": valid_row["map_id"],
        "game_mode": valid_row["game_mode"],
        "game_type": valid_row["game_type"],
        "game_version": valid_row["game_version"],
        "game_start_timestamp": valid_row["game_start_timestamp"],
        "game_end_timestamp": valid_row["game_end_timestamp"],
        "game_start_local": valid_row["game_start_local"],
        "game_end_local": valid_row["game_end_local"],
        "game_duration_seconds": valid_row["game_duration_seconds"],
        "player_subteam_id": valid_row["player_subteam_id"],
        "subteam_placement": valid_row["subteam_placement"],
        "points": valid_row["points"],
        "candidate_source": valid_row["candidate_source"],
        "match_json_path": valid_row["match_json_path"],
        "timeline_json_path": valid_row["timeline_json_path"],
        "p1_champion_name": p1.get("champion_name"),
        "p1_champion_id": p1.get("champion_id"),
        "p1_kills": p1.get("kills"),
        "p1_deaths": p1.get("deaths"),
        "p1_assists": p1.get("assists"),
        "p1_gold_earned": p1.get("gold_earned"),
        "p1_total_damage_dealt_to_champions": p1.get("total_damage_dealt_to_champions"),
        "p1_total_damage_taken": p1.get("total_damage_taken"),
        "p1_player_augment1": p1.get("player_augment1"),
        "p1_player_augment2": p1.get("player_augment2"),
        "p1_player_augment3": p1.get("player_augment3"),
        "p1_player_augment4": p1.get("player_augment4"),
        "p2_champion_name": p2.get("champion_name"),
        "p2_champion_id": p2.get("champion_id"),
        "p2_kills": p2.get("kills"),
        "p2_deaths": p2.get("deaths"),
        "p2_assists": p2.get("assists"),
        "p2_gold_earned": p2.get("gold_earned"),
        "p2_total_damage_dealt_to_champions": p2.get("total_damage_dealt_to_champions"),
        "p2_total_damage_taken": p2.get("total_damage_taken"),
        "p2_player_augment1": p2.get("player_augment1"),
        "p2_player_augment2": p2.get("player_augment2"),
        "p2_player_augment3": p2.get("player_augment3"),
        "p2_player_augment4": p2.get("player_augment4"),
    }


def load_or_download_match(
    client: RiotClient,
    regional_cluster: str,
    match_id: str,
    raw_matches_dir: Path,
    raw_timelines_dir: Path,
    include_timeline: bool,
    force: bool,
) -> Tuple[Path, Optional[Path], Dict[str, Any]]:
    ensure_dir(raw_matches_dir)
    if include_timeline:
        ensure_dir(raw_timelines_dir)

    match_path = raw_matches_dir / f"{match_id}.json"
    timeline_path = raw_timelines_dir / f"{match_id}_timeline.json" if include_timeline else None

    if match_path.exists() and not force:
        match_payload = load_json(match_path)
    else:
        match_payload = client.get_match(regional_cluster, match_id)
        save_json(match_path, match_payload)

    if include_timeline and timeline_path is not None:
        if not timeline_path.exists() or force:
            timeline_payload = client.get_timeline(regional_cluster, match_id)
            save_json(timeline_path, timeline_payload)

    return match_path, timeline_path, match_payload


def match_start_epoch_seconds(match_payload: Dict[str, Any]) -> Optional[int]:
    info = match_payload.get("info", {})
    value = info.get("gameStartTimestamp")
    if value is None:
        return None
    return int(value / 1000)


def evaluate_duo_match(
    duo: DuoRef,
    match_payload: Dict[str, Any],
    window: TournamentWindow,
    match_path: Path,
    timeline_path: Optional[Path],
    source_label: str,
) -> Tuple[bool, Dict[str, Any]]:
    info = match_payload.get("info", {})
    metadata = match_payload.get("metadata", {})
    match_id = metadata.get("matchId")

    base = {
        "duo_id": duo.duo_id,
        "duo_name": duo.duo_name,
        "player1_riot_id": duo.player1_riot_id,
        "player2_riot_id": duo.player2_riot_id,
        "player1_puuid": duo.player1_puuid,
        "player2_puuid": duo.player2_puuid,
        "platform": duo.platform,
        "regional_cluster": duo.regional_cluster,
        "match_id": match_id,
        "queue_id": info.get("queueId"),
        "game_mode": info.get("gameMode"),
        "game_start_timestamp": info.get("gameStartTimestamp"),
        "game_start_local": epoch_ms_to_local_iso(info.get("gameStartTimestamp")),
        "candidate_source": source_label,
        "match_json_path": str(match_path.resolve()),
        "timeline_json_path": str(timeline_path.resolve()) if timeline_path else None,
    }

    queue_id = info.get("queueId")
    if queue_id not in ARENA_QUEUES:
        base["reason"] = "not_arena_queue"
        return False, base

    game_start_seconds = match_start_epoch_seconds(match_payload)
    if game_start_seconds is None:
        base["reason"] = "missing_game_start_timestamp"
        return False, base

    if game_start_seconds < window.start_epoch_seconds or game_start_seconds > window.end_epoch_seconds:
        base["reason"] = "outside_tournament_window"
        return False, base

    p1 = find_participant(match_payload, duo.player1_puuid)
    p2 = find_participant(match_payload, duo.player2_puuid)

    if p1 is None and p2 is None:
        base["reason"] = "both_players_missing_from_match"
        return False, base
    if p1 is None:
        base["reason"] = "player1_missing_from_match"
        return False, base
    if p2 is None:
        base["reason"] = "player2_missing_from_match"
        return False, base

    subteam1 = p1.get("playerSubteamId")
    subteam2 = p2.get("playerSubteamId")

    if subteam1 is None or subteam2 is None:
        base["reason"] = "missing_player_subteam_id"
        base["player1_subteam_id"] = subteam1
        base["player2_subteam_id"] = subteam2
        return False, base

    if subteam1 != subteam2:
        base["reason"] = "players_not_in_same_subteam"
        base["player1_subteam_id"] = subteam1
        base["player2_subteam_id"] = subteam2
        return False, base

    placement1 = p1.get("subteamPlacement")
    placement2 = p2.get("subteamPlacement")
    if placement1 is None or placement2 is None:
        base["reason"] = "missing_subteam_placement"
        return False, base

    if placement1 != placement2:
        base["reason"] = "placement_mismatch_between_teammates"
        base["player1_subteam_placement"] = placement1
        base["player2_subteam_placement"] = placement2
        return False, base

    points = points_from_placement(placement1)
    if points is None:
        base["reason"] = "unsupported_subteam_placement"
        base["subteam_placement"] = placement1
        return False, base

    return True, build_valid_row(
        duo=duo,
        match_payload=match_payload,
        timeline_path=timeline_path,
        match_path=match_path,
        source_label=source_label,
    )


def build_scoreboard_rows(valid_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    scoreboard: Dict[str, Dict[str, Any]] = {}

    for row in valid_rows:
        key = row["duo_id"]
        entry = scoreboard.setdefault(key, {
            "duo_id": row["duo_id"],
            "duo_name": row["duo_name"],
            "player1_riot_id": row["player1_riot_id"],
            "player2_riot_id": row["player2_riot_id"],
            "valid_matches": 0,
            "total_points": 0,
            "placements_1": 0,
            "placements_2": 0,
            "placements_3": 0,
            "placements_4": 0,
            "placements_5": 0,
            "placements_6": 0,
            "placements_7": 0,
            "placements_8": 0,
        })

        placement = row.get("subteam_placement")
        points = row.get("points") or 0

        entry["valid_matches"] += 1
        entry["total_points"] += points

        if placement in range(1, 9):
            entry[f"placements_{placement}"] += 1

    return sorted(
        scoreboard.values(),
        key=lambda x: (-x["total_points"], -x["placements_1"], -x["placements_2"], x["duo_id"])
    )


def parse_args() -> argparse.Namespace:
    window = default_window()

    parser = argparse.ArgumentParser(
        description="Programa 1 do campeonato Arena: coleta, valida e pontua partidas válidas de duplas."
    )
    parser.add_argument("--api-key", default=os.getenv("RIOT_API_KEY"), help="Chave da Riot API. Também aceita RIOT_API_KEY.")
    parser.add_argument("--duos-xlsx", required=True, help="Arquivo XLSX com as colunas duo_id, player1_riot_id, player2_riot_id, duo_name.")
    parser.add_argument("--output", default="./arena_program1_dump", help="Diretório de saída.")
    parser.add_argument("--start-time", type=int, default=window.start_epoch_seconds, help="Epoch seconds inicial. Default: início do campeonato.")
    parser.add_argument("--end-time", type=int, default=window.end_epoch_seconds, help="Epoch seconds final. Default: fim do campeonato.")
    parser.add_argument("--no-timeline", action="store_true", help="Não baixa timeline.")
    parser.add_argument("--max-matches-per-queue", type=int, help="Limite por queue para teste. O padrão é sem limite.")
    parser.add_argument("--force", action="store_true", help="Rebaixa arquivos de partida já existentes e invalida o cache de jogadores.")
    parser.add_argument("--verbose", action="store_true", help="Ativa logs detalhados.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    if not args.api_key:
        print("Erro: informe --api-key ou defina RIOT_API_KEY.", file=sys.stderr)
        return 2

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
    )

    run_started_monotonic = time.time()

    output_dir = Path(args.output)
    ensure_dir(output_dir)

    window = TournamentWindow(
        start_local_iso=datetime.fromtimestamp(args.start_time, tz=TOURNAMENT_TZ).isoformat(),
        end_local_iso=datetime.fromtimestamp(args.end_time, tz=TOURNAMENT_TZ).isoformat(),
        start_epoch_seconds=args.start_time,
        end_epoch_seconds=args.end_time,
    )

    try:
        duos_raw = load_duos_xlsx(Path(args.duos_xlsx))
    except Exception as exc:
        print(f"Erro ao carregar duplas.xlsx: {exc}", file=sys.stderr)
        return 2

    client = RiotClient(api_key=args.api_key)

    # 1) Resolver jogadores únicos
    unique_riot_ids: List[str] = []
    for row in duos_raw:
        unique_riot_ids.append(row["player1_riot_id"])
        unique_riot_ids.append(row["player2_riot_id"])
    unique_riot_ids = dedupe_keep_order(unique_riot_ids)

    # Cache: Riot IDs já resolvidos em execuções anteriores não precisam ser
    # re-resolvidos (PUUID e plataforma não mudam). --force invalida o cache.
    if args.force:
        cached_players: Dict[str, PlayerRef] = {}
        logging.info("--force: ignorando cache de players_resolved.json.")
    else:
        cached_players = load_players_cache(output_dir / "players_resolved.json")
        if cached_players:
            logging.info("Cache de players carregado: %d entradas.", len(cached_players))

    players: Dict[str, PlayerRef] = {}
    player_errors: List[Dict[str, Any]] = []

    for riot_id in unique_riot_ids:
        cached = cached_players.get(riot_id)
        if cached is not None:
            players[riot_id] = cached
            logging.debug("Player %s veio do cache (skip API).", riot_id)
            continue

        try:
            logging.info("Resolvendo conta de %s", riot_id)
            game_name, tag_line = split_riot_id(riot_id)
            account = client.resolve_account_by_riot_id(game_name, tag_line)
            puuid = account["puuid"]
            platform = client.get_active_platform(puuid)
            regional_cluster = platform_to_regional_cluster(platform)
            players[riot_id] = PlayerRef(
                riot_id=riot_id,
                game_name=game_name,
                tag_line=tag_line,
                puuid=puuid,
                platform=platform,
                regional_cluster=regional_cluster,
                slug=slugify(riot_id),
            )
        except Exception as exc:
            logging.exception("Falha ao resolver %s", riot_id)
            player_errors.append({
                "riot_id": riot_id,
                "error": str(exc),
            })

    save_json(output_dir / "players_resolved.json", {
        "players": [asdict(player) for player in players.values()],
        "errors": player_errors,
    })

    # 2) Montar duplas resolvidas
    duos: List[DuoRef] = []
    duo_resolution_errors: List[Dict[str, Any]] = []

    for row in duos_raw:
        p1 = players.get(row["player1_riot_id"])
        p2 = players.get(row["player2_riot_id"])

        if p1 is None or p2 is None:
            duo_resolution_errors.append({
                **row,
                "error": "player_resolution_failed",
            })
            continue

        if p1.platform != p2.platform:
            duo_resolution_errors.append({
                **row,
                "error": "players_on_different_platforms",
                "player1_platform": p1.platform,
                "player2_platform": p2.platform,
            })
            continue

        if p1.regional_cluster != p2.regional_cluster:
            duo_resolution_errors.append({
                **row,
                "error": "players_on_different_regional_clusters",
                "player1_regional_cluster": p1.regional_cluster,
                "player2_regional_cluster": p2.regional_cluster,
            })
            continue

        duos.append(DuoRef(
            duo_id=row["duo_id"],
            duo_name=row["duo_name"],
            player1_riot_id=row["player1_riot_id"],
            player2_riot_id=row["player2_riot_id"],
            player1_puuid=p1.puuid,
            player2_puuid=p2.puuid,
            platform=p1.platform,
            regional_cluster=p1.regional_cluster,
            slug=slugify(f'{row["duo_id"]}_{row["duo_name"]}'),
        ))

    save_json(output_dir / "duos_resolved.json", {
        "window": asdict(window),
        "duos": [asdict(duo) for duo in duos],
        "errors": duo_resolution_errors,
    })

    # 3) Coletar match IDs por jogador na janela
    player_match_ids: Dict[str, Dict[str, Any]] = {}
    player_match_id_errors: List[Dict[str, Any]] = []

    for riot_id, player in players.items():
        try:
            all_match_ids: List[str] = []
            queue_breakdown: Dict[str, int] = {}
            for queue_id in ARENA_QUEUES:
                logging.info("Buscando match IDs de %s | queue=%s", riot_id, queue_id)
                ids = client.get_match_ids(
                    regional_cluster=player.regional_cluster,
                    puuid=player.puuid,
                    queue_id=queue_id,
                    start_time=window.start_epoch_seconds,
                    end_time=window.end_epoch_seconds,
                    max_matches=args.max_matches_per_queue,
                )
                queue_breakdown[str(queue_id)] = len(ids)
                all_match_ids.extend(ids)

            all_match_ids = dedupe_keep_order(all_match_ids)
            player_match_ids[riot_id] = {
                "riot_id": riot_id,
                "puuid": player.puuid,
                "platform": player.platform,
                "regional_cluster": player.regional_cluster,
                "queue_breakdown": queue_breakdown,
                "total_unique_match_ids": len(all_match_ids),
                "match_ids": all_match_ids,
            }
        except Exception as exc:
            logging.exception("Falha ao coletar match IDs de %s", riot_id)
            player_match_id_errors.append({
                "riot_id": riot_id,
                "error": str(exc),
            })

    save_json(output_dir / "player_match_ids.json", {
        "window": asdict(window),
        "players": list(player_match_ids.values()),
        "errors": player_match_id_errors,
    })

    # 4) Baixar/validar partidas por dupla
    raw_matches_dir = output_dir / "raw" / "matches"
    raw_timelines_dir = output_dir / "raw" / "timelines"

    valid_rows: List[Dict[str, Any]] = []
    invalid_rows: List[Dict[str, Any]] = []
    download_errors: List[Dict[str, Any]] = []
    downloaded_match_ids: Set[str] = set()

    for duo in duos:
        p1_ids = set(player_match_ids.get(duo.player1_riot_id, {}).get("match_ids", []))
        p2_ids = set(player_match_ids.get(duo.player2_riot_id, {}).get("match_ids", []))
        candidate_ids = dedupe_keep_order(list(p1_ids & p2_ids))

        logging.info(
            "Validando dupla %s (%s) | candidatos=%s",
            duo.duo_id, duo.duo_name, len(candidate_ids)
        )

        for idx, match_id in enumerate(candidate_ids, start=1):
            source_parts: List[str] = []
            if match_id in p1_ids:
                source_parts.append("player1")
            if match_id in p2_ids:
                source_parts.append("player2")
            source_label = "+".join(source_parts) if source_parts else "unknown"

            try:
                logging.info(
                    "[%s/%s] Dupla %s | baixando/validando %s",
                    idx, len(candidate_ids), duo.duo_id, match_id
                )

                match_path, timeline_path, match_payload = load_or_download_match(
                    client=client,
                    regional_cluster=duo.regional_cluster,
                    match_id=match_id,
                    raw_matches_dir=raw_matches_dir,
                    raw_timelines_dir=raw_timelines_dir,
                    include_timeline=not args.no_timeline,
                    force=args.force,
                )
                downloaded_match_ids.add(match_id)

                is_valid, payload = evaluate_duo_match(
                    duo=duo,
                    match_payload=match_payload,
                    window=window,
                    match_path=match_path,
                    timeline_path=timeline_path,
                    source_label=source_label,
                )
                if is_valid:
                    valid_rows.append(payload)
                else:
                    invalid_rows.append(payload)

            except Exception as exc:
                logging.exception("Falha na dupla %s para a partida %s", duo.duo_id, match_id)
                download_errors.append({
                    "duo_id": duo.duo_id,
                    "duo_name": duo.duo_name,
                    "match_id": match_id,
                    "error": str(exc),
                })

    # 5) Deduplicar válidas por dupla+match_id
    unique_valid: List[Dict[str, Any]] = []
    seen_valid_keys: Set[Tuple[str, str]] = set()
    for row in valid_rows:
        key = (row["duo_id"], row["match_id"])
        if key not in seen_valid_keys:
            seen_valid_keys.add(key)
            unique_valid.append(row)
    valid_rows = unique_valid

    # 6) Persistência
    save_json(output_dir / "valid_matches.json", {
        "window": asdict(window),
        "count": len(valid_rows),
        "rows": valid_rows,
    })
    save_json(output_dir / "invalid_matches.json", {
        "window": asdict(window),
        "count": len(invalid_rows),
        "rows": invalid_rows,
    })
    save_json(output_dir / "download_errors.json", {
        "count": len(download_errors),
        "rows": download_errors,
    })

    valid_csv_rows = [build_valid_csv_row(row) for row in valid_rows]
    flatten_csv_rows(output_dir / "duo_matches.csv", valid_csv_rows)

    scoreboard_rows = build_scoreboard_rows(valid_rows)
    flatten_csv_rows(output_dir / "scoreboard.csv", scoreboard_rows)

    run_id = datetime.now(TOURNAMENT_TZ).isoformat()
    duration_seconds = round(time.time() - run_started_monotonic, 2)

    public_snapshot = {
        "schema_version": 1,
        "run_id": run_id,
        "generated_at": run_id,
        "collector_duration_seconds": duration_seconds,
        "window": asdict(window),
        "scoreboard": scoreboard_rows,
        "duos_resolved": [asdict(duo) for duo in duos],
        "players_resolved": [asdict(player) for player in players.values()],
        "valid_matches": valid_rows,
        "counts": {
            "duos_input_count": len(duos_raw),
            "duos_resolved_count": len(duos),
            "players_unique_count": len(unique_riot_ids),
            "players_resolved_count": len(players),
            "valid_matches_count": len(valid_rows),
            "invalid_matches_count": len(invalid_rows),
            "download_errors_count": len(download_errors),
        },
    }
    save_json(output_dir / "public_snapshot.json", public_snapshot)

    run_manifest = {
        "schema_version": 1,
        "run_id": run_id,
        "generated_at": run_id,
        "duration_seconds": duration_seconds,
        "window": asdict(window),
        "duos_input_count": len(duos_raw),
        "duos_resolved_count": len(duos),
        "players_unique_count": len(unique_riot_ids),
        "players_resolved_count": len(players),
        "raw_unique_matches_downloaded_or_loaded": len(downloaded_match_ids),
        "valid_matches_count": len(valid_rows),
        "invalid_matches_count": len(invalid_rows),
        "player_resolution_errors_count": len(player_errors),
        "duo_resolution_errors_count": len(duo_resolution_errors),
        "player_match_id_errors_count": len(player_match_id_errors),
        "download_errors_count": len(download_errors),
        "include_timeline": not args.no_timeline,
        "output_dir": str(output_dir.resolve()),
        "paths": {
            "players_resolved_json": str((output_dir / "players_resolved.json").resolve()),
            "duos_resolved_json": str((output_dir / "duos_resolved.json").resolve()),
            "player_match_ids_json": str((output_dir / "player_match_ids.json").resolve()),
            "valid_matches_json": str((output_dir / "valid_matches.json").resolve()),
            "invalid_matches_json": str((output_dir / "invalid_matches.json").resolve()),
            "duo_matches_csv": str((output_dir / "duo_matches.csv").resolve()),
            "scoreboard_csv": str((output_dir / "scoreboard.csv").resolve()),
            "public_snapshot_json": str((output_dir / "public_snapshot.json").resolve()),
            "raw_matches_dir": str(raw_matches_dir.resolve()),
            "raw_timelines_dir": str(raw_timelines_dir.resolve()) if not args.no_timeline else None,
        },
    }
    save_json(output_dir / "run_manifest.json", run_manifest)

    logging.info(
        "Finalizado. Duplas resolvidas=%s | válidas=%s | inválidas=%s | erros download=%s",
        len(duos), len(valid_rows), len(invalid_rows), len(download_errors)
    )

    return 0 if not download_errors else 1


if __name__ == "__main__":
    raise SystemExit(main())